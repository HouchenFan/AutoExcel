import numpy as np
import os
import sys
import time
import re
# from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font


# 获取某列的所有值
def get_col_value(ws, column):
    rows = ws.max_row
    column_data = []
    for i in range(1, rows + 1):
        cell_value = ws.cell(row=i, column=column).value
        column_data.append(cell_value)
    return column_data

def Scoring(days):
    if days >= 30:
        score = 25
    elif days >= 15 and days < 30:
        score = 20
    elif days < 15:
        score = 12.5
    else:
        score = 0
    return score

def filter(str):
    return int(str.strip('转正').split('提前')[-1])

class Table():
    def __int__(self, wb, str):
        self.ws_raw = wb[str]
        self.ws_grp_rate = wb['最佳组长打分表']
        self.ws_prt_rate = wb['最佳科长打分表']
        self.grp_name_list = [] # 各组名称列表
        self.prt_name_list = [] # 各科名称列表
        for x in range(7, self.ws_grp_rate.max_row):
            cell_value = self.ws_grp_rate.cell(row=x, column=1).value
            if cell_value:
                self.grp_name_list.append(cell_value)
        for x in range(7, self.ws_prt_rate.max_row):
            cell_value = self.ws_prt_rate.cell(row=x, column=1).value
            if cell_value:
                self.prt_name_list.append(cell_value)

    def GrpCpltRate(self):
        raw_col_list = get_col_value(self.ws_raw, 1) # 季度原始表格第一列数据
        grp_name_idx = []
        # 获取季度原始表中各组组名的index
        for name in self.grp_name_list:
            if name:
                idx = raw_col_list.index(name)
                grp_name_idx.append(idx + 1)
        # 判断两表的小组是否一致
        if len(self.grp_name_list) != len(grp_name_idx):
            print('Error: 小组数量对不上！！！')
            sys.exit(0)
        # 找到各组对应的数据， 并进行复制
        self.grp_act_days_list = []
        self.grp_act_days_rate_list = []
        for i in range(len(grp_name_idx)):
            key_word = 'members joined more than 6 months'
            row_grp_name = grp_name_idx[i]
            row_raw = raw_col_list[row_grp_name:row_grp_name + 100].index(key_word) + row_grp_name + 1
            col_raw_1 = column_index_from_string('AL')
            col_raw_2 = column_index_from_string('AM')
            row_grp = i + 7
            col_grp = column_index_from_string('G')
            cell_value_1 = self.ws_raw.cell(row=row_raw, column=col_raw_1).value
            cell_value_2 = self.ws_raw.cell(row=row_raw, column=col_raw_2).value
            self.grp_act_days_list.append(cell_value_1)
            self.grp_act_days_rate_list.append(cell_value_2)
            self.ws_grp_rate.cell(row=row_grp, column=col_grp).value = cell_value_2


    def GrpMembQualRate(self):
        col_raw = column_index_from_string('L')
        col_new = column_index_from_string('C')
        raw_col_list = get_col_value(self.ws_grp_rate, col_raw)
        re1 = r'提前(.*?)天'
        font = Font(u'微软雅黑', size=10, bold=True, italic=False, strike=False, color='FF0000')
        for i in range(7, self.ws_grp_rate.max_row):
            cell_value_raw = raw_col_list[i-1]
            try:
                result = re.findall(re1, cell_value_raw)
                if '平均' in cell_value_raw:
                    self.ws_grp_rate.cell(row=i, column=col_new).font = font
                    score = 99999
                else:
                    mean_day = np.mean(list(map(filter, result)))
                    score = Scoring(mean_day)
            except:
                pass
            self.ws_grp_rate.cell(row=i, column=col_new).value = score
            score = None

    def PrtMembQualRate(self):
        col_raw = column_index_from_string('L')
        col_new = column_index_from_string('C')
        raw_col_list = get_col_value(self.ws_prt_rate, col_raw)
        re1 = r'提前(.*?)天'
        font = Font(u'微软雅黑', size=10, bold=True, italic=False, strike=False, color='FF0000')
        for i in range(7, self.ws_prt_rate.max_row):
            cell_value_raw = raw_col_list[i-1]
            try:
                result = re.findall(re1, cell_value_raw)
                if '平均' in cell_value_raw:
                    self.ws_prt_rate.cell(row=i, column=col_new).font = font
                    score = 99999
                else:
                    mean_day = np.mean(list(map(filter, result)))
                    score = Scoring(mean_day)
            except:
                pass
            self.ws_prt_rate.cell(row=i, column=col_new).value = score
            score = None

    def PrtCplRate(self):
        self.grp_of_prt_idx_list = []
        self.grp_cpl_mounts = np.array([a*b for (a, b) in zip(self.grp_act_days_list, self.grp_act_days_rate_list)])
        for prt_idx, prt_name in enumerate(self.prt_name_list):
            grp_idx_list = []
            if prt_name:
                for grp_idx, grp_name in enumerate(self.grp_name_list):
                    if grp_name:
                        if prt_name in grp_name:
                            grp_idx_list.append(grp_idx)

            prt_cpl_rate = np.sum(self.grp_cpl_mounts[grp_idx_list]) / np.sum(np.array(self.grp_act_days_list)[grp_idx_list])
            self.ws_prt_rate.cell(row=prt_idx + 7, column=column_index_from_string('G')).value = prt_cpl_rate



if __name__ == '__main__':
    new_table_path = r"/Users/fanhouchen/Desktop/表格/NewTable.xlsx"  # 新表格的路径
    quarter = 'Q4'
    # 读取表格
    print("开始读取表格！")
    wb_new = load_workbook(new_table_path, data_only=True)
    print("表格读取完成！")

    wb = Table()
    wb.__int__(wb_new, quarter)

    t = time.perf_counter()  # 记录当前时间

    wb.GrpCpltRate()  # 复制各组已转正员工季度平均完成率（CN)
    wb.GrpMembQualRate() # 组长的组员转正达标分数
    wb.PrtMembQualRate() # 科长的组员转正达标分数
    wb.PrtCplRate() # 计算各科已转正员工季度平均完成率


    print(f'coast:{time.perf_counter() - t:.8f}s')  # 打印总时间

    wb_new.save(r"/Users/fanhouchen/Desktop/表格/NewTable_out.xlsx")
    wb_new.close()
